import vk_api
import requests
import re
from datetime import datetime, timedelta
import os
import urllib.request
import subprocess
import csv
import camelot
import sys
import time
from threading import Thread
from vk_api import VkApi
from vk_api.bot_longpoll import VkBotLongPoll, VkBotEventType
import AdflyAPI
from fuzzywuzzy import fuzz
from openpyxl import load_workbook

#const--
admin_id=[230245992]
file = open("groups.txt", "r")
i=0
groups_id=[]
groups_name=[]
for line in file.readlines():
    if i==0:
        groups_id.append(int(line))
        i+=1
    else:
        i=0
        groups_name.append(int(line))
file.close()
months1=['',
        'январь',
        'февраль',
        'март',
        'апрель',
        'май',
        'июнь',
        'июль',
        'август',
        'сентябрь',
        'октябрь',
        'ноябрь',
        'декабрь']


#--const

#function for sending messages array of id
def message_send(vk, peers, text):
    #vk - get_api()
    #peers - array of id(int)
    #text - message
    try:
        for _id in peers:
            print("send to: ",_id)
            vk.messages.send(random_id=0, peer_id= _id ,message=text)
    except:
        pass

#find id in array of id
def compare_id(_id, array_id):
    a=False
    for i in array_id:
        if _id == i:
            a=True
    return a
class global_sending(Thread):
    def __init__(self, vk, id):
        Thread._
        _init__(self)
        self.id=id
    def run(self):
        vk.messages.send(random_id=0, peer_id= self.id ,message="Напишите номера групп.")
        f=1
        for event in longpoll.listen():
            if event.type == VkBotEventType.MESSAGE_NEW:
                if f==1 and event.object['peer_id']==self.id:
                    t=re.findall('[0-9]{3}',event.object['text'].lower())
                    vk.messages.send(random_id=0, peer_id= self.id ,message="Отправим: "+", ".join(t))
                    vk.messages.send(random_id=0, peer_id= self.id ,message="Напишите ваше сообщение.")
                    f=2
                elif f==2 and event.object['peer_id']==self.id:
                    vk.messages.send(random_id=0, peer_id= self.id ,message="Я отправлю это сообщение:")
                    vk.messages.send(random_id=0, peer_id= self.id ,message=event.object['text'])
                    text=event.object['text']
                    vk.messages.send(random_id=0, peer_id= self.id ,message="Вы уверены?(Д, Н)")
                    f=3
                elif f==3 and event.object['peer_id']==self.id:
                    if event.object['text']=='Д':
                        t_id=[]
                        for i in t:
                            try:
                                t_id.append(groups_id[groups_name.index(int(i))])
                            except:
                                vk.messages.send(random_id=0, peer_id= self.id ,message="group "+i+"not found")
                        message_send(vk, t_id, text)
                
                    
                    

                    
                    
class monitor(Thread):
    def __init__(self, vk):
        Thread.__init__(self)
    def run(self):
        f=1
        while 1:
            if f==0:
                if findfile(datetime.now() + timedelta(days=1))[0] == 0:
                    message_send(vk, groups_id, '@all !!ВСЕ СЮДА!! На сайте появилось долгожданное расписание!! Спросите меня и я вам расскажу какое у вас.')
                    #Прошу прощение  за беспокойство, ща всё исправим))
                    f=1
            elif datetime.now().strftime("%I%M")=='0010':
                f=0
            time.sleep(50)
class MyThread(Thread):
    #распараллеливание
    #новый поток сообщений 
    #передаём id и 'vk'
    def __init__(self, vk, id):
        Thread.__init__(self)
        self.id = id
        self.vk = vk
    
    def run(self):
        f=0
        if groups_id.count(self.id)==0 and f==0:
            vk.messages.send(random_id=0, peer_id=self.id,message='Я забыл из какой ты группы, напиши номер группы пж')
            f=1
        else:
            vk.messages.send(random_id=0, peer_id=self.id,message='На когда группа '+str(groups_name[groups_id.index(self.id)])+' желает расписание? p.s. Если требуется исправить номер группы обращайтесь в личку сообщества, вам обязательно помогут.')
            f=2
        for event in longpoll.listen():
            if f==1 and re.fullmatch('[0-9]{3}',event.object['text'].lower()) and event.object['peer_id']==self.id:
                groups_name.append(re.search('[0-9]{3}',event.object['text'].lower()).group())
                groups_id.append(self.id)
                file=open('groups.txt', 'a')
                file.write(str(self.id)+'\n')
                file.write(re.search('[0-9]{3}',event.object['text'].lower()).group()+'\n')
                file.close()
                vk.messages.send(random_id=0, peer_id=self.id,message='На когда группа '+str(groups_name[groups_id.index(self.id)])+' желает расписание?')
                f=2
            elif f==1:
                vk.messages.send(random_id=0, peer_id=self.id,message='Номер группы пожалуйста..')
            elif re.fullmatch('^(после)*завтра',event.object['text'].lower()) and f==2 and self.id==event.object['peer_id']:
                date=datetime.now() + timedelta(days=1+len(re.findall('после',event.object['text'].lower())))
                date = date.strftime("%d.%m.%y")
                vk.messages.send(random_id=0, peer_id=self.id,message='Ща буит расписание на '+date)
                f=3
            elif re.fullmatch('^сегодня',event.object['text'].lower()) and f==2 and self.id==event.object['peer_id']:
                date=datetime.now().strftime("%d.%m.%y")
                vk.messages.send(random_id=0, peer_id=self.id,message='Ща буит расписание на '+date)
                f=3
            elif re.fullmatch('[0-9]{2}\.[0-9]{2}\.[0-9]{2}',event.object['text'].lower()) and f==2 and self.id==event.object['peer_id']:
                date=re.search('[0-9]{2}\.[0-9]{2}\.[0-9]{2}',event.object['text'].lower()).group()
                vk.messages.send(random_id=0, peer_id=event.object['peer_id'],message='Ща буит расписание на '+date)
                f=3
            elif event.object['peer_id']==self.id and f==2:
                vk.messages.send(random_id=0, peer_id=self.id,message='Напиши, когда? Завтра? Сегодня? Полслезавтра?')
            if f==3:
                go(self.id, datetime.strptime(date, "%d.%m.%y"), str(groups_name[groups_id.index(self.id)]))
                return
#starting find, creating and sending list
def go(id, date, group):
    url = findfile(date)
    print('url', url)
    if url[0]==1:
        vk.messages.send(random_id=0, peer_id=id,message="Расписания нет(((")
    else:
        
        rasp = parserXls(url[1], group)
        file = open(r"txt/message.txt", "w")
        file.write("Расписание на "+date.strftime("%d.%m.%y") + '\n')
        for i in range(len(rasp)):
            for j in range(len(rasp[i])):
                file.write(str(rasp[i][j])+'\n')
        file.close()
        file = open(r"txt/message.txt", "r")
        vk.messages.send(random_id=0, peer_id=id,message=file.read())
        file.close()

def getCsv(name):#function from NET
    tables = camelot.read_pdf('./'+name+'.pdf')
    tables
    # <TableList n=1>
    tables.export('foo.csv', f='csv', compress=True) # json, excel, html, sqlite
    tables[0]
    # <Table shape=(7, 7)>
    tables[0].parsing_report
    {
        'accuracy': 99.02,
        'whitespace': 12.24,
        'order': 1,
        'page': 1
    }
    tables[0].to_csv('csv/'+name+'.csv') # to_json, to_excel, to_html, to_sqlite
    tables[0].df # get a pandas DataFrame!
def get_url_adfly(surname):
#     getCsv('lectors')
    mass=[]
    with open('csv/lectors.csv', newline='') as File:  
        reader = csv.reader(File)
        for row in reader:
#             print(fuzz.ratio(row[1], surname)>90, row[1], surname)
            if fuzz.partial_ratio(row[1], surname)>70:
                message_send(vk, admin_id, str(fuzz.ratio(row[1], surname))+' '+ str(row[1])+' '+ str(surname))
#                 return AdflyAPI.convert_to_adfly(row[2])
                return row[2]
def find_group_in_xl(wb, group):
    rasp=[]
    for sheet in wb.sheetnames:
#         print(1)
        sheet = wb[sheet]
        i=4
#         while sheet.cell(row=1, column=i).value!=None:
        while i<100:
#             print(2, sheet.cell(row=1, column=i).value, str(sheet.cell(row=1, column=i).value).find(str(group)) , group)
            if str(sheet.cell(row=1, column=i).value).find(str(group))!=-1:
                j=2
#                 print('========================================================================')
                while sheet.cell(row=j, column=i).value==None:
                    j+=2
                while sheet.cell(row=j, column=i).value!=None:
                    if fuzz.ratio(str(sheet.cell(row=j, column=i+1).value), "Спортзал")>80:
                        f="===== sport class"
                    elif sheet.cell(row=j+1, column=i+1).value == None:
                        u=get_url_adfly(sheet.cell(row=j+1, column=i).value)
                        if u==None:
                            u=get_url_adfly(sheet.cell(row=j, column=i).value)
                        if u==None:
                            f="online. По поводу ссылки обращайтесь к преподавателю."
                        else:
                            f="online: "+u
                    else:
                        f="===== "+str(sheet.cell(row=j+1, column=i+1).value)
                    if sheet.cell(row=j+1, column=i).value==None:
                        rasp.append([str(int(j/2))+' пара', sheet.cell(row=j, column=i).value, f])
                    else:
                        rasp.append([str(int(j/2))+' пара', sheet.cell(row=j, column=i).value,sheet.cell(row=j+1, column=i).value, f])
#                     rasp.append([str(int(j/2))+' пара', sheet.cell(row=j, column=i).value,p, f])
                    j+=2
            i+=2
    return rasp

def parserXls(file_name, group):
#     print("libreoffice -convert-to xlsx " + file_name + " " + file_name + "x")
    os.system("libreoffice --convert-to xlsx " + file_name + " --outdir ./pdf")
    file_name+="x"
    wb = load_workbook(file_name)
    return find_group_in_xl(wb, group)
    
# def createRasp(date, group, rasp):
#     file=open(r"txt/message.txt", 'w')
#     return 'Rasp na '+ date.strftime("%d.%m.%y") + '\n group '+ group+ '\n'.join(rasp)

def getUrl(line):
    i=0
    res=''
    while(i<len(line)):
        if line[i]=='<' and line[i+1]=='a' and line[i+2]==' ' and line[i+3]=='h' and line[i+4]=='r' and line[i+5]=='e' and line[i+6]=='f' and line[i+7]=='=':
            i+=9
            
            while(line[i]!='\"'):
                res+=line[i]
                i+=1
                if line[i]=='&' and line[i+1]=='a' and line[i+2]=='m' and line[i+3]=='p':
                    break
        
        i+=1
    return res
def findfile(now):
    now2=now.strftime("%d.%m.%y")
    now=str(now.day)+' '+months1[now.month]
    
    # try:
    #     if open(now+'.pdf'):
    #         res=0
    # except FileNotFoundError:
    # # else: <a href="/stud
    link=urllib.request.urlopen('http://www.mgkit.ru/studentu/raspisanie-zanatij')
    #go to url and open html file
    lines = []
    for line in link.readlines():
        line=line.decode('utf-8')
        print(fuzz.partial_ratio(now, line))
        print(now)
        print(line)
        if fuzz.partial_ratio(now, line)>88: # find files now date
            lines.append(re.findall(r"<a href=..st[^>][^>]*",line)) #save strings with links
#     print(lines)
    for i in range(len(lines)):
        lines[i]=getUrl(lines[i][0]) # save oll lonks#
    link.close()
#     print(lines)
    if len(lines)==0:
        res=1
    else:
        os.system('wget -O ./pdf/'+now2+'.xls '+'http://www.mgkit.ru'+lines[0]+' ') # get file
        res=0
    return [res, './pdf/'+now2+'.xls']

def create_monitoring(vk):
    thread = monitor(vk)
    thread.start()
def create_thread(vk, id):
    thread = MyThread(vk, id)
    thread.start()
def global_send(vk, id):
    thread = global_sending(vk, id)
    thread.start()
while 1:
    try:
        vk_session = VkApi(token="14ddcb3b6b0b32b3eb6267d173aff130fa57d50869ce8aae629a96cd9ef749651c9d5e1705fd7e1683458")
        longpoll = VkBotLongPoll(vk_session, "194668032")
        vk = vk_session.get_api()

        message_send(vk, admin_id, "I'm work")
        create_monitoring(vk)

        for event in longpoll.listen():
            try:
                if event.object['peer_id']>2000000000:
                    print(groups_name[groups_id.index(event.object['peer_id'])], event.object['text'])
            except:
                pass
            if event.type == VkBotEventType.MESSAGE_NEW:
                if re.fullmatch('^.расписание', event.object['text'].lower()):
                    create_thread(vk, event.object['peer_id'])
                elif compare_id(event.object['peer_id'], admin_id) and re.fullmatch(r'^.рассылка.*', event.object['text'].lower()):
                    print('maill')
                    global_send(vk, event.object['peer_id'])
                    # mailing(event.object['text'].lower())

                elif event.object['peer_id']<2000000000:
                    print(event.object['peer_id'])
                    # vk.messages.send(random_id=0, peer_id= 230245992 ,message="from "+"https://vk.com/id"+str(event.object['from_id'])+'\n'+str(event.object['text']))
    except Exception as e:
        message_send(vk, admin_id, e)
